#!/usr/bin/python3
# -*- coding:utf-8 -*-
# For Python 3.6+

from TaikoTools.Taiko_dat import StoryDatFile, SysDatFile, CharaDatFile
from openpyxl import Workbook,load_workbook
from bs4 import BeautifulSoup
from TaikoTools.consts import *
import time, os

class SysDatExcel(SysDatFile):
    def __init__(self, table_name):
        self.table_name = table_name
        super(SysDatExcel, self).__init__("")
        pass
    def import_from_table(self):
        wb = load_workbook(self.table_name)
        for file in sys_dat_files:
            super(SysDatExcel, self).__init__(os.path.join(sys_dat_file_dir, file))
            self.dat_read()
            try:
                ws = wb.get_sheet_by_name(self.title.decode(encoding = "utf-8"))
            except:
                ws = wb.get_sheet_by_name("√" + self.title.decode(encoding = "utf-8"))
            #print(self.contents)
            for row in ws.iter_rows(min_row=2, max_col=3, max_row=self.text_count+1):
                if row[2].value == None: continue
                #print(row[0].value, row[2].value)
                self.contents[row[0].value.encode()] = str(row[2].value).encode()
            self.dat_write()
            print("[Table]: %s is Done." % ws.title)
    
    def export_to_table(self):
        wb = Workbook()
        for file in sys_dat_files:
            super(SysDatExcel, self).__init__(os.path.join(sys_dat_file_dir, file))
            self.dat_read()
            ws = wb.create_sheet(self.title.decode(encoding = "utf-8"))
            ws.cell(row = 1, column = 1).value = "标识"
            ws.cell(row = 1, column = 2).value = "原文"
            ws.cell(row = 1, column = 3).value = "译文"
            for i, (flag, txt) in enumerate(self.contents.items()):
                ws.cell(row = i + 2, column = 1).value = flag.decode(encoding = "utf-8")
                ws.cell(row = i + 2, column = 2).value = txt.decode(encoding = "utf-8")
            print("[Table]: %s is Done." % ws.title)
        wb.save(self.table_name)
        
    def export_trans_to_table(self):
        wb = load_workbook(self.table_name)
        for file in sys_dat_files:
            super(SysDatExcel, self).__init__(os.path.join(sys_dat_file_dir, file + "_tr"))
            self.dat_read()
            try:
                ws = wb.get_sheet_by_name(self.title.decode(encoding = "utf-8"))
            except:
                ws = wb.get_sheet_by_name("√" + self.title.decode(encoding = "utf-8"))
            for row in ws.iter_rows(min_row=2, max_col=3, max_row=self.text_count+1):
                try:
                    row[2].value = self.contents[row[0].value.encode()]
                except KeyError:
                    print("[Error]: Source_tr_file %s losed flag!" % (file + "_tr"))
            print("[Table]: %s is Done." % ws.title)
        wb.save(self.table_name)

class CharaDatExcel(CharaDatFile):
    def __init__(self, table_name):
        self.table_name = table_name
        super(CharaDatExcel, self).__init__("")
        pass
    def import_from_table(self):
        wb = load_workbook(self.table_name)
        for file in chara_dat_files:
            super(CharaDatExcel, self).__init__(os.path.join(chara_dat_file_dir, file))
            self.dat_read()
            try:
                ws = wb.get_sheet_by_name(file)
            except:
                ws = wb.get_sheet_by_name("√" + file)
            #print(self.contents)
            for row in ws.iter_rows(min_row=2, max_col=3, max_row=self.text_count+1):
                if row[2].value == None: continue
                #print(row[0].value, row[2].value)
                self.contents[row[0].value.encode()]['txt'] = str(row[2].value).encode()
            self.dat_write()
            print("[Table]: %s is Done." % ws.title)
    
    def export_to_table(self):
        wb = Workbook()
        for file in chara_dat_files:
            super(CharaDatExcel, self).__init__(os.path.join(chara_dat_file_dir, file))
            self.dat_read()
            ws = wb.create_sheet(file)
            ws.cell(row = 1, column = 1).value = "标识"
            ws.cell(row = 1, column = 2).value = "原文"
            ws.cell(row = 1, column = 3).value = "译文"
            for i, (flag, content) in enumerate(self.contents.items()):
                ws.cell(row = i + 2, column = 1).value = flag.decode(encoding = "utf-8")
                ws.cell(row = i + 2, column = 2).value = content['txt'].decode(encoding = "utf-8")
            print("[Table]: %s is Done." % ws.title)
        wb.save(self.table_name)
        
    def export_trans_to_table(self):
        wb = load_workbook(self.table_name)
        for file in chara_dat_files:
            super(CharaDatExcel, self).__init__(os.path.join(chara_dat_file_dir, file + "_tr"))
            self.dat_read()
            try:
                ws = wb.get_sheet_by_name(file)
            except:
                ws = wb.get_sheet_by_name("√" + file)
            for row in ws.iter_rows(min_row=2, max_col=3, max_row=self.text_count+1):
                try:
                    row[2].value = self.contents[row[0].value.encode()]['txt']
                except KeyError:
                    print("[Error]: Source_tr_file %s losed flag!" % (file + "_tr"))
            print("[Table]: %s is Done." % ws.title)
        wb.save(self.table_name)

class StoryDatExcel(StoryDatFile):
    def __init__(self, table_name):
        self.table_name = table_name
        super(StoryDatExcel, self).__init__("")
        pass
        
    def txt_import_process(self, flag, txt):
        self.contents[flag]['phonetic'] = [] # Don't need phonetic char
        soup = BeautifulSoup(txt, "html.parser")
        #print(list(soup.strings))
        if len(list(soup.strings)) != len(self.contents[flag]['txt'][0]): raise TypeError
        txt = []
        for s in soup.strings:
            tag = s.parent.name
            if tag == 'color-blue': 
                txt.append([s.encode(), COLOR_BLUE])
            elif tag == 'color-unknown': 
                txt.append([s.encode(), COLOR_UNKNOWN])
            elif s != '': txt.append([s.encode(), COLOR_NULL])
        self.contents[flag]['txt'][0] = txt
    
    def txt_export_process(self, txt):
        txt_formated = ''
        for (txt_raw , txt_color) in txt:
                color_tag = COLOR_MAP[txt_color]
                txt_formated += color_tag[0] + txt_raw.decode(encoding = "utf-8") + color_tag[1]
        return txt_formated        
    
    def import_from_table(self):
        wb = load_workbook(self.table_name)
        for file in story_dat_files:
            super(StoryDatExcel, self).__init__(os.path.join(story_dat_file_dir, file))
            self.dat_read()
            try:
                ws = wb.get_sheet_by_name(file)
            except:
                ws = wb.get_sheet_by_name("√" + file)
            #print(self.contents)
            for row in ws.iter_rows(min_row=2, max_col=3, max_row=self.text_count+1):
                if row[2].value == None: continue
                #print(row[0].value, row[2].value)
                self.txt_import_process(row[0].value.encode(), row[2].value)
            print("[Table]: %s is Done." % ws.title)
            self.dat_write()
    
    def export_to_table(self):
        wb = Workbook()
        for file in story_dat_files:
            super(StoryDatExcel, self).__init__(os.path.join(story_dat_file_dir, file))
            self.dat_read()
            ws = wb.create_sheet(file)
            ws.cell(row = 1, column = 1).value = "标识"
            ws.cell(row = 1, column = 2).value = "原文"
            ws.cell(row = 1, column = 3).value = "译文"
            for i, (flag, content) in enumerate(self.contents.items()):
                ws.cell(row = i + 2, column = 1).value = flag.decode(encoding = "utf-8")
                #print(flag, content)
                ws.cell(row = i + 2, column = 2).value = self.txt_export_process(content['txt'][0])
            print("[Table]: %s is Done." % ws.title)
        wb.save(self.table_name)
    
    def export_trans_to_table(self):
        wb = load_workbook(self.table_name)
        for file in story_dat_files:
            super(StoryDatExcel, self).__init__(os.path.join(story_dat_file_dir, file + "_tr"))
            self.dat_read()
            try:
                ws = wb.get_sheet_by_name(file)
            except:
                ws = wb.get_sheet_by_name("√" + file)
            for row in ws.iter_rows(min_row=2, max_col=3, max_row=self.text_count+1):
                try:
                    row[2].value = self.txt_export_process(self.contents[row[0].value.encode()]['txt'][0])
                except KeyError:
                    print("[Error]: Source_tr_file %s losed flag!" % (file + "_tr"))
            print("[Table]: %s is Done." % ws.title)
        wb.save(self.table_name)
    
if __name__ == "__main__":
    #SysTable = SysDatExcel("sys_text.xlsx")
    #SysTable.import_from_table()
    #SysTable.export_to_table()
    #SysTable.export_trans_to_table()
    
    StoryTable = StoryDatExcel("story_text.xlsx")
    StoryTable.import_from_table()
    #StoryTable.export_to_table()
    #StoryTable.export_trans_to_table()
    
    #CharaTable = CharaDatExcel("chara_text.xlsx")
    #CharaTable.import_from_table()
    #CharaTable.export_to_table()
    #CharaTable.export_trans_to_table()
    